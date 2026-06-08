"""
Dump all products from Nahdi, United Pharmacy, and Al-Dawaa into an Excel file.
Each pharmacy gets its own sheet. Uses parallel fetching for speed.
"""
import requests, cloudscraper, json, time, re, sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "tawfeery_all_products.xlsx"


def parse_nahdi_item(item):
    sku = item.get('sku', '')
    name = item.get('name', '')
    price = float(item.get('price', 0))
    img_path = ''
    qty = None
    unit_desc = ''
    for attr in item.get('custom_attributes', []):
        code = attr.get('attribute_code', '')
        val = attr.get('value', '')
        if code == 'image':
            img_path = val
        elif code == 'quantity_and_unit_description':
            unit_desc = val
            m = re.search(r'(\d+)', val)
            if m:
                qty = int(m.group(1))
        elif code == 'size':
            m = re.search(r'(\d+)', str(val))
            if m:
                qty = int(m.group(1))
    image_url = ''
    if img_path:
        image_url = 'https://ecombe.nahdionline.com/media/catalog/product' + img_path
    return {
        'SKU': sku,
        'Name (AR)': name,
        'Name (EN)': name,
        'Price (SAR)': price if price else '',
        'Quantity': qty or '',
        'Unit': unit_desc,
        'Image URL': image_url,
        'Product Link': 'https://www.nahdionline.com/ar-sa/search?query=' + name.replace(' ', '+'),
        'Brand': '',
        'Offer': '',
    }


def fetch_nahdi_page(scraper, page, page_size):
    params = {
        'searchCriteria[filterGroups][0][filters][0][field]': 'name',
        'searchCriteria[filterGroups][0][filters][0][value]': '%%',
        'searchCriteria[filterGroups][0][filters][0][conditionType]': 'like',
        'searchCriteria[pageSize]': str(page_size),
        'searchCriteria[currentPage]': str(page),
    }
    res = scraper.get('https://ecombe.nahdionline.com/rest/V1/products',
                      params=params, timeout=30, headers={'Accept': 'application/json'})
    res.raise_for_status()
    return res.json()


def fetch_nahdi():
    print("[Nahdi] Starting dump...")
    scraper = cloudscraper.create_scraper()
    page_size = 1000

    # First page to get total
    data = fetch_nahdi_page(scraper, 1, page_size)
    total = data.get('total_count', 0)
    total_pages = (total + page_size - 1) // page_size
    print("[Nahdi] Total: %d products, %d pages" % (total, total_pages))

    all_products = [parse_nahdi_item(item) for item in data.get('items', [])]
    print("[Nahdi] Page 1: %d/%d" % (len(all_products), total))

    def fetch_page(page_num):
        try:
            d = fetch_nahdi_page(scraper, page_num, page_size)
            return page_num, [parse_nahdi_item(item) for item in d.get('items', [])]
        except Exception as e:
            print("[Nahdi] Error page %d: %s" % (page_num, e))
            return page_num, None

    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(fetch_page, p): p for p in range(2, total_pages + 1)}
        done_count = 1
        for future in as_completed(futures):
            page_num, items = future.result()
            if items is not None:
                all_products.extend(items)
            done_count += 1
            if done_count % 50 == 0 or done_count == total_pages:
                print("[Nahdi] Progress: %d/%d pages, %d products" % (done_count, total_pages, len(all_products)))

    print("[Nahdi] Done. %d products." % len(all_products))
    return all_products


def fetch_united():
    print("[United] Starting dump...")
    all_products = []
    headers = {
        'X-Algolia-API-Key': 'NGFkYzM5MDgzYjA0YmI2YzdlYjk4YjIwNDFjZjQzZTg2ZDQ4M2Q0ZGM5ZTVjYTgxYTNjZWRlMjllZDg0YTg3Y3RhZ0ZpbHRlcnM9',
        'X-Algolia-Application-Id': 'Y1GOQ9DTV8'
    }
    url = 'https://Y1GOQ9DTV8-dsn.algolia.net/1/indexes/*/queries'
    hits_per_page = 1000
    page = 0

    while True:
        try:
            payload = {
                "requests": [{
                    "indexName": "unitedpharmacy_livear_products",
                    "params": "query=&hitsPerPage=%d&page=%d" % (hits_per_page, page)
                }]
            }
            res = requests.post(url, headers=headers, json=payload, timeout=20)
            res.raise_for_status()
            result = res.json().get('results', [{}])[0]

            if page == 0:
                total = result.get('nbHits', 0)
                nb_pages = result.get('nbPages', 0)
                print("[United] Total: %d products, %d pages" % (total, nb_pages))

            hits = result.get('hits', [])
            if not hits:
                break

            for h in hits:
                name = h.get('name', '')
                price_val = h.get('price', 0)
                if isinstance(price_val, dict):
                    price_val = price_val.get('SAR', {}).get('default', 0)
                elif isinstance(price_val, list):
                    price_val = price_val[0] if price_val else 0
                img_url = h.get('image_url') or h.get('thumbnail_url', '')
                link = h.get('url', '')
                sku = h.get('objectID') or h.get('sku') or ''
                brand = h.get('brand', '')
                name_en = h.get('name_locally_en', '')
                offer = ''
                if h.get('isOfferApplicable'):
                    offer = h.get('offerApplicableLabel', '')
                all_products.append({
                    'SKU': sku,
                    'Name (AR)': name,
                    'Name (EN)': name_en,
                    'Price (SAR)': float(price_val) if price_val else '',
                    'Quantity': '',
                    'Unit': '',
                    'Image URL': img_url,
                    'Product Link': link,
                    'Brand': brand,
                    'Offer': offer,
                })

            if page % 5 == 0:
                print("[United] Page %d: %d products so far" % (page, len(all_products)))

            if page >= result.get('nbPages', 0) - 1:
                break
            page += 1

        except Exception as e:
            print("[United] Error page %d: %s" % (page, e))
            time.sleep(2)

    print("[United] Done. %d products." % len(all_products))
    return all_products


def parse_aldawaa_item(p):
    name = p.get('name', '')
    if not name:
        return None
    price_val = p.get('price', {}).get('value')
    if price_val is None:
        return None
    img_url = ''
    image_urls = p.get('imageUrl', [])
    if isinstance(image_urls, list) and len(image_urls) > 0:
        ar_img = next((img.get('value') for img in image_urls if img.get('key') == 'ar'), None)
        img_url = ar_img or image_urls[0].get('value', '')
    if img_url and img_url.startswith('/'):
        img_url = 'https://stgprevapi.al-dawaa.com' + img_url
    link = p.get('url', '')
    if link and not link.startswith('http'):
        link = 'https://www.al-dawaa.com' + link
    offer_text = ''
    potential = p.get('potentialPromotions', [])
    if isinstance(potential, list) and len(potential) > 0:
        promo_code = potential[0].get('code', '').strip()
        if promo_code and 'توصيل' not in promo_code:
            offer_text = promo_code
    sku = p.get('code', '')
    brand = ''
    brand_data = p.get('brand')
    if isinstance(brand_data, dict):
        brand = brand_data.get('name', '') or ''
    name_en = p.get('urlProductName', '') or ''
    return {
        'SKU': sku,
        'Name (AR)': name,
        'Name (EN)': name_en,
        'Price (SAR)': float(price_val),
        'Quantity': '',
        'Unit': '',
        'Image URL': img_url,
        'Product Link': link,
        'Brand': brand,
        'Offer': offer_text,
    }


def fetch_aldawaa_page(page, page_size):
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept-Language': 'ar,en-US;q=0.9,en;q=0.8',
    }
    params = {
        'query': '',
        'pageSize': page_size,
        'currentPage': page,
        'lang': 'ar',
        'curr': 'SAR'
    }
    res = requests.get('https://stgprevapi.al-dawaa.com/occ/v2/aldawaa/products/search',
                       params=params, headers=headers, timeout=20)
    res.raise_for_status()
    return res.json()


def fetch_aldawaa():
    print("[Al-Dawaa] Starting dump...")
    all_products = []
    page_size = 100

    data = fetch_aldawaa_page(0, page_size)
    total_pages = data.get('pagination', {}).get('totalPages', 0)
    print("[Al-Dawaa] Total pages: %d (~%d products)" % (total_pages, total_pages * page_size))

    def parse_page_products(data):
        results = []
        for p in data.get('products', []):
            item = parse_aldawaa_item(p)
            if item:
                results.append(item)
        return results

    all_products.extend(parse_page_products(data))
    print("[Al-Dawaa] Page 0: %d products" % len(all_products))

    def fetch_page(page_num):
        try:
            d = fetch_aldawaa_page(page_num, page_size)
            return page_num, parse_page_products(d)
        except Exception as e:
            print("[Al-Dawaa] Error page %d: %s" % (page_num, e))
            return page_num, None

    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(fetch_page, p): p for p in range(1, total_pages)}
        done_count = 1
        for future in as_completed(futures):
            page_num, items = future.result()
            if items is not None:
                all_products.extend(items)
            done_count += 1
            if done_count % 50 == 0 or done_count >= total_pages - 1:
                print("[Al-Dawaa] Progress: %d/%d pages, %d products" % (done_count, total_pages, len(all_products)))

    print("[Al-Dawaa] Done. %d products." % len(all_products))
    return all_products


def write_excel(nahdi, united, aldawaa, output):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    columns = ['SKU', 'Name (AR)', 'Name (EN)', 'Price (SAR)', 'Quantity', 'Unit', 'Image URL', 'Product Link', 'Brand', 'Offer']
    col_widths = [20, 50, 50, 14, 12, 30, 50, 60, 30, 40]

    datasets = [
        ('Nahdi Online', nahdi),
        ('United Pharmacy', united),
        ('Al-Dawaa', aldawaa),
    ]
    first_sheet = True
    for sheet_name, data in datasets:
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx, product in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = product.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_name in ('Price (SAR)', 'Quantity') and val != '':
                    cell.alignment = Alignment(horizontal='right')
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(columns)), len(data) + 1)
        print("Sheet '%s': %d rows" % (sheet_name, len(data)))
    wb.save(output)
    print("\nSaved: %s" % output)


if __name__ == '__main__':
    t0 = time.time()
    nahdi = fetch_nahdi()
    united = fetch_united()
    aldawaa = fetch_aldawaa()
    write_excel(nahdi, united, aldawaa, OUTPUT)
    elapsed = time.time() - t0
    total = len(nahdi) + len(united) + len(aldawaa)
    print("\nTotal: %d products in %ds" % (total, elapsed))
