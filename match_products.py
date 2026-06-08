"""
Smart Product Matching Engine v5
- Parses product names into structured fields (brand, product, variant, size)
- Uses reference database for exact matching
- Suggests alternatives when exact match not available
- Outputs unified catalog with alternatives
"""
import xlrd, re, time, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = r'C:\Users\POS\Desktop\tawfeery\tawfeery_all_products.xls'
OUTPUT = r'C:\Users\POS\Desktop\tawfeery\tawfeery_matched_products.xlsx'


# ── Smart Product Parser ─────────────────────────────────────────────────────

# Known brands (expandable) - first word(s) of product name
KNOWN_BRANDS = {
    'nivea', 'dove', 'pantene', 'garnier', 'loreal', "nature's", 'natures',
    'maybelline', 'rimmel', 'bourjois', 'essence', 'flormar', 'max',
    'palmers', 'johnson', 'dabur', 'rexona', 'ogx', 'beesline', 'avene',
    'bioderma', 'eucerin', 'la', 'cetaphil', 'cerave', 'vichy', 'clinique',
    'neutrogena', 'olay', 'samsung', 'philips', 'always', 'baby', 'joy',
    'marnys', 'redoxon', 'caltrate', 'centrum', 'solaray', 'carlson',
    'holland', 'barrett', 'enspire', 'ensur', 'ensure', ' aptamil',
    'aptamil', 'nan', 'nestle', 'domol', 'pampers', 'chicco', 'huggies',
    'dettol', 'lysol', 'dettol', 'tcp', 'savlon', 'listine', 'colgate',
    'oral', 'sensodyne', 'signal', 'closeup', 'pepsodent', 'meridol',
    'vanish', 'tide', 'ariel', 'persil', 'omo', 'downy', 'comfort',
    'dettol', 'airwick', 'glade', 'febreze', 'lynx', 'axe', 'old', 'spice',
    'adidas', 'nike', 'reebok', 'puma', 'calvin', 'klein', 'davidoff',
    'chanel', 'dior', 'gucci', 'versace', 'armani', 'prada', 'boss',
    'dove', 'lux', 'lifebuoy', 'shield', 'sure', 'degree', 'speedstick',
    'gillette', 'braun', 'panasonic', 'sony', 'lg', 'bosch', 'siemens',
    'tommee', 'tippee', 'dr', 'brown', 'avent', 'mam', 'hegen',
    'lansinoh', 'medela', 'spectrum', 'neutrogena', 'cetaphil', 'cerave',
    'la', 'roche', 'posay', 'bioderma', 'avene', 'uriage', 'evoluderm',
    'aDerma', 'ducray', 'klorane', 'weleda', 'boiron', 'homeolab',
    'nuby', 'nuk', 'baby', 'björn', 'belen', 'babyjoy', 'baby', 'joy',
    'aquamarine', 'jP', 'glovit', 'nemr', 'royal', 'apisal', 'avogain',
    'super', 'prune', 'beauty', 'form', 'volum', 'herbal', 'essences',
    'tRESemmé', 'tresseme', 'tressemmé', 'organics', 'alpecin', 'kérastase',
    'wella', 'schwarzkopf', 'dove', 'men', 'clear', 'head', 'shoulders',
    'sunsilk', '顿', '顿', '乐', '康', '维', '善', '存', '善存',
    'caltrate', 'centrum', 'redoxon', 'berocca', 'emetrol', 'dramamine',
    'tylenol', 'advil', 'motrin', 'aleve', 'bayer', 'aspirin', 'tylenol',
    'panadol', 'nurofen', 'ibuprofen', 'paracetamol', 'acetaminophen',
    'voltaren', 'diclofenac', 'cataflam', 'arcoxia', 'etoricoxib',
    'nexium', 'omeprazole', 'pantoprazole', 'ranitidine', 'zantac',
    'gaviscon', 'maalox', 'mylanta', 'pepto', 'bismol', 'imodium',
    'loperamide', 'dulcolax', 'bisacodyl', 'senokot', 'senna',
    'claritin', 'loratadine', 'zyrtec', 'cetirizine', 'allegra', 'fexofenadine',
    'beconase', 'flonase', 'nasacort', 'rhinocort', 'nasal',
    'cough', 'delsym', 'robitussin', 'benylin', 'dimetapp', 'vicks',
    'nyquil', 'dayquil', 'thalion', 'theraflu', 'fervex', 'humex',
    ' augmentin', 'amoxicillin', 'azithromycin', 'cipro', 'ciprofloxacin',
    'keflex', 'cephalexin', 'bactrim', 'trimethoprim', 'nitrofurantoin',
    'fluconazole', 'nystatin', 'clotrimazole', 'terbinafine', 'lamisil',
    'diflucan', 'sporanox', 'itraconazole', 'valtrex', 'acyclovir',
    'zovirax', 'tamiflu', 'oseltamivir', 'relenza', 'zanamivir',
    'insulin', 'metformin', 'glucophage', 'januvia', 'sitagliptin',
    'galvus', 'vildagliptin', 'glyxambi', 'empagliflozin', 'jardiance',
    'synthroid', 'levothyroxine', 'euthyrox', 'thyrox', 'cytomel', 'liothyronine',
    'lipitor', 'atorvastatin', 'crestor', 'rosuvastatin', 'zocor', 'simvastatin',
    'plavix', 'clopidogrel', 'brilinta', 'ticagrelor', 'effient', 'prasugrel',
    'coumadin', 'warfarin', 'xarelto', 'rivaroxaban', 'eliquis', 'apixaban',
    'pradaxa', 'dabigatran', 'safil', 'aspirin', 'ecotrin', 'low', 'dose',
    'norvasc', 'amlodipine', 'diovan', 'valsartan', 'cozaar', 'losartan',
    'accupril', 'quinapril', 'lotensin', 'benazepril', 'mavik', 'trandolapril',
    'toprol', 'metoprolol', 'lopressor', 'tenormin', 'atenolol', 'coreg', 'carvedilol',
    'lanoxin', 'digoxin', 'pacerone', 'amiodarone', 'cardizem', 'diltiazem',
    'verelan', 'verapamil', 'calan', 'procardia', 'nifedipine', 'adalat',
    'nitrostat', 'nitroglycerin', 'imdur', 'isosorbide', 'isordil',
    'lasix', 'furosemide', 'bumex', 'bumetanide', 'demadex', 'torsemide',
    'hydrodiuril', 'hydrochlorothiazide', 'microzide', 'hctz',
    'aldactone', 'spironolactone', 'inspra', 'eplerenone',
    'prednisone', 'deltasone', 'methylprednisolone', 'medrol',
    'prednisolone', 'orapred', 'pediapred', 'ketorolac', 'toradol',
    'tramadol', 'ultram', 'conzip', 'nucynta', 'tapentadol',
    'oxycontin', 'oxycodone', 'percocet', 'norco', 'vicodin', 'hydrocodone',
    'dilaudid', 'hydromorphone', 'fentanyl', 'duragesic', 'subsys',
    'morphine', 'ms', 'contin', 'kadian', 'embeda',
    'gabapentin', 'neurontin', 'pregabalin', 'lyrica',
    'topamax', 'topiramate', 'depakote', 'divalproex', 'valproic',
    'tegretol', 'carbamazepine', 'trileptal', 'oxcarbazepine',
    'lamictal', 'lamotrigine', 'keppra', 'levetiracetam', 'dilantin', 'phenytoin',
    'zonegran', 'zonisamide', 'felbatol', 'felbamate', 'luminal', 'phenobarbital',
    'sabril', 'vigabatrin', 'lyrica', 'pregabalin', 'gabapentin', 'neurontin',
    'celexa', 'citalopram', 'lexapro', 'escitalopram', 'prozac', 'fluoxetine',
    'paxil', 'paroxetine', 'zoloft', 'sertraline', 'effexor', 'venlafaxine',
    'cymbalta', 'duloxetine', 'pristiq', 'desvenlafaxine', 'fetzima', 'levomilnacipran',
    'wellbutrin', 'bupropion', 'zyban', 'forfivo',
    'remeron', 'mirtazapine', 'trazodone', 'desyrel',
    'sinequan', 'doxepin', 'norpramin', 'desipramine', 'tofranil', 'imipramine',
    'anafranil', 'clomipramine', 'pamelor', 'nortriptyline', 'aviptyl', 'amitriptyline',
    'elavil', 'endep', 'sinequan', 'doxepin', 'surmontil', 'trimipramine',
    'stelazine', 'trifluoperazine', 'thorazine', 'chlorpromazine', 'compazine', 'prochlorperazine',
    'invega', 'paliperidone', 'risperdal', 'risperidone', 'abilify', 'aripiprazole',
    'seroquel', 'quetiapine', 'zyprexa', 'olanzapine', 'geodon', 'ziprasidone',
    'latuda', 'lurasidone', 'saphris', 'asenapine', 'fanapt', 'iloperidone',
    'risperdal', 'consta', 'invega', 'sustenna', 'invega', 'trinza',
    'haldol', 'haloperidol', 'decanoate', 'navane', 'thiothixene',
    'loxapine', 'loxitane', 'molindone', 'moban', 'pelaperidone',
    'cogentin', 'benztropine', 'artane', 'trihexyphenidyl',
    'lithium', 'eskalith', 'lithobid', 'depakote', 'divalproex',
    'lamictal', 'lamotrigine', 'tegretol', 'carbamazepine',
    'tramadol', 'ultram', 'tylenol', 'acetaminophen', 'advil', 'ibuprofen',
    'aleve', 'naproxen', 'aspirin', 'bayer', 'excedrin',
    'pepto', 'bismol', 'maalox', 'mylanta', 'tums', 'ralox',
    'gaviscon', 'prilosec', 'omeprazole', 'nexium', 'esomeprazole',
    'protonix', 'pantoprazole', 'prevacid', 'lansoprazole',
    'aciphex', 'rabeprazole', 'dexilant', 'dexlansoprazole',
    'tagamet', 'cimetidine', 'pepcid', 'famotidine', 'zantac', 'ranitidine',
    'imodium', 'loperamide', 'dulcolax', 'bisacodyl', 'senokot', 'senna',
    'metamucil', 'psyllium', 'fibercon', 'polycarbophil', 'miralax', 'polyethylene',
    'milk', 'magnesia', ' Phillips',
}

# Product form keywords
FORM_WORDS = {
    'tablet', 'tab', 'cap', 'capsule', 'caplet', 'syrup', 'suspension',
    'susp', 'solution', 'sol', 'drops', 'drop', 'cream', 'ointment',
    'gel', 'lotion', 'spray', 'inhaler', 'neb', 'nebulizer', 'inhalation',
    'powder', 'granule', 'sachet', 'effervescent', 'eff', 'chewable',
    'softgel', 'gelcap', 'liquid', 'suspension', 'emulsion', 'lotion',
    'shampoo', 'conditioner', 'serum', 'ampoule', 'amp', 'injection',
    'inj', 'suppository', 'pessary', 'patch', 'strip', 'film',
    'bar', 'soap', 'deodorant', 'deo', 'toothpaste', 'mouthwash',
    'diaper', 'culotte', 'pant', 'wipe', 'tissue', 'pad',
}

# Size/quantity patterns
SIZE_RE = re.compile(
    r'(\d+(?:\.\d+)?)\s*(mg|mcg|g|gm|ml|l|iu|pcs?|tabs?|caps?|caplets?|'
    r'tablets?|capsules?|sachets?|strips?|packs?|boxes?|bottles?|tubes?|'
    r'sheets?|pads?|pcs?|pieces?|units?|doses?|injections?|ampoules?|amp?|'
    r'softgels?|gelcaps?|effervescent|drops?|sprays?|nebulizer|inhaler)',
    re.IGNORECASE
)

# Variant keywords (for distinguishing product variants)
VARIANT_KEYWORDS = {
    'extra', 'forte', 'plus', 'max', 'super', 'ultra', 'pro', 'pro-v',
    'intensive', 'sensitive', 'dry', 'oily', 'normal', 'combination',
    'mature', 'young', 'baby', 'kids', 'junior', 'men', 'women',
    'light', 'medium', 'dark', 'fair', 'white', 'natural', 'original',
    'classic', 'premium', 'gold', 'silver', 'platinum', 'diamond',
    'soft', 'extra', 'cotton', 'clean', 'fresh', 'pure', 'gentle',
    'hypoallergenic', 'organic', 'natural', 'herbal', 'botanical',
    'anti', 'anti-dandruff', 'anti-hair', 'anti-acne', 'anti-aging',
    'vitamin', 'mineral', 'omega', 'calcium', 'iron', 'zinc',
    'with', 'without', 'free', 'plus', ' enriched',
}


def parse_product_name(name, brand_hint=''):
    """Parse product name into structured fields."""
    if not name:
        return None

    original = str(name).strip()
    # Al-Dawaa uses hyphens
    is_aladawaa = '-' in original and ' ' not in original
    if is_aladawaa:
        normalized = original.replace('-', ' ')
    else:
        normalized = original

    lower = normalized.lower().strip()

    # Extract size/quantity
    size_match = SIZE_RE.search(lower)
    size_val = ''
    size_unit = ''
    if size_match:
        size_val = size_match.group(1)
        size_unit = size_match.group(2).rstrip('s')  # normalize plural

    # Extract brand from English name only (brand column is unreliable - empty or Arabic)
    brand = ''
    tokens = lower.split()
    if tokens:
        # Check if first word(s) form a known brand
        for n_words in [2, 1]:
            candidate = ' '.join(tokens[:n_words])
            if candidate in KNOWN_BRANDS:
                brand = candidate
                break
        if not brand:
            brand = tokens[0]

    # Extract form
    form = ''
    for fw in FORM_WORDS:
        if fw in lower:
            form = fw
            break

    # Extract variant keywords
    variants = []
    for vk in VARIANT_KEYWORDS:
        if vk in lower:
            variants.append(vk)
    variant_str = ' '.join(sorted(set(variants)))

    # Build product key (brand + product name without size)
    # Remove size pattern from name for product identification
    product_key = SIZE_RE.sub('', lower).strip()
    # Remove brand from product key
    if brand and product_key.startswith(brand):
        product_key = product_key[len(brand):].strip()

    return {
        'original': original,
        'brand': brand,
        'product_key': product_key,
        'size_val': size_val,
        'size_unit': size_unit,
        'form': form,
        'variant': variant_str,
        'normalized': lower,
    }


# ── Matching Logic ───────────────────────────────────────────────────────────

def exact_match_score(parsed_a, parsed_b):
    """Score for identical product (same brand, product, size)."""
    if not parsed_a or not parsed_b:
        return 0.0

    # Brand must match (or both have same first token)
    if parsed_a['brand'] != parsed_b['brand']:
        # Allow if brands share first token
        ba = parsed_a['brand'].split()
        bb = parsed_b['brand'].split()
        if not (ba and bb and ba[0] == bb[0]):
            return 0.0

    # Product key similarity
    tokens_a = set(parsed_a['product_key'].split())
    tokens_b = set(parsed_b['product_key'].split())
    if not tokens_a or not tokens_b:
        return 0.0

    inter = tokens_a & tokens_b
    union = tokens_a | tokens_b
    jaccard = len(inter) / len(union) if union else 0.0

    # Size must match
    size_match = (parsed_a['size_val'] == parsed_b['size_val'] and
                  parsed_a['size_unit'] == parsed_b['size_unit'])

    # Form should match
    form_match = (parsed_a['form'] == parsed_b['form']) or (not parsed_a['form'] or not parsed_b['form'])

    if jaccard >= 0.5 and size_match and form_match:
        return 0.8 + 0.2 * jaccard

    return 0.0


def alternative_match_score(parsed_a, parsed_b):
    """Score for alternative product (same brand, different variant/size)."""
    if not parsed_a or not parsed_b:
        return 0.0

    if parsed_a['brand'] != parsed_b['brand']:
        return 0.0

    tokens_a = set(parsed_a['product_key'].split())
    tokens_b = set(parsed_b['product_key'].split())
    if not tokens_a or not tokens_b:
        return 0.0

    inter = tokens_a & tokens_b
    union_set = tokens_a | tokens_b
    jaccard = len(inter) / len(union_set) if union_set else 0.0

    if jaccard >= 0.3:
        return 0.5 + 0.3 * jaccard

    return 0.0


# ── Load Data ────────────────────────────────────────────────────────────────

def load_sheet(wb, name):
    ws = wb.sheet_by_name(name)
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(1, ws.nrows):
        row = {}
        for c in range(ws.ncols):
            val = ws.cell_value(r, c)
            if isinstance(val, float) and val == int(val):
                val = int(val)
            row[headers[c]] = val
        rows.append(row)
    return rows


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    t0 = time.time()
    print("Loading data...")
    wb = xlrd.open_workbook(INPUT)
    nahdi = load_sheet(wb, 'Nahdi Online')
    united = load_sheet(wb, 'United Pharmacy')
    aldawaa = load_sheet(wb, 'Al-Dawaa')
    wb.release_resources()

    # Parse all products
    print("Parsing product names...")
    all_products = []
    for store_name, data in [('Nahdi', nahdi), ('United', united), ('Al-Dawaa', aldawaa)]:
        for p in data:
            name = str(p.get('Name (EN)', ''))
            if not name:
                continue
            parsed = parse_product_name(name, str(p.get('Brand', '')))
            if parsed:
                all_products.append({
                    'store': store_name,
                    'parsed': parsed,
                    'name': name,
                    'name_ar': str(p.get('Name (AR)', '')),
                    'price': p.get('Price (SAR)', ''),
                    'sku': str(p.get('SKU', '')),
                    'link': str(p.get('Product Link', '')),
                    'img': str(p.get('Image URL', '')),
                })

    print("Total parsed: %d" % len(all_products))

    # Build brand index for fast lookup
    print("Building brand index...")
    brand_index = defaultdict(list)
    for i, p in enumerate(all_products):
        brand_index[p['parsed']['brand']].append(i)

    # ── Matching ──────────────────────────────────────────────────────────────
    print("\nMatching identical products...")
    n = len(all_products)

    # Union-Find for exact matches
    parent = list(range(n))
    rank = [0] * n

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx == ry:
            return
        if rank[rx] < rank[ry]:
            rx, ry = ry, rx
        parent[ry] = rx
        if rank[rx] == rank[ry]:
            rank[rx] += 1

    exact_pairs = 0
    processed = 0

    # For each brand, compare products within that brand
    for brand, indices in brand_index.items():
        if len(indices) < 2:
            continue

        # Compare all pairs within this brand
        for ii in range(len(indices)):
            i = indices[ii]
            pi = all_products[i]['parsed']
            for jj in range(ii + 1, len(indices)):
                j = indices[jj]
                pj = all_products[j]['parsed']

                # Must be different stores
                if all_products[i]['store'] == all_products[j]['store']:
                    continue

                score = exact_match_score(pi, pj)
                if score >= 0.7:
                    union(i, j)
                    exact_pairs += 1

        processed += len(indices)
        if processed % 10000 == 0:
            print("  %d/%d products processed, %d exact pairs" % (processed, n, exact_pairs))

    print("Exact match pairs: %d" % exact_pairs)

    # ── Find alternatives (same brand, not exact match) ───────────────────────
    print("\nFinding alternatives...")
    alternatives = defaultdict(list)  # key: group_root, value: list of (product_idx, score)

    # For alternatives, compare products from different stores with same brand
    for brand, indices in brand_index.items():
        if len(indices) < 2:
            continue
        for ii in range(len(indices)):
            i = indices[ii]
            pi = all_products[i]['parsed']
            root_i = find(i)
            for jj in range(ii + 1, len(indices)):
                j = indices[jj]
                pj = all_products[j]['parsed']

                if all_products[i]['store'] == all_products[j]['store']:
                    continue

                # If not in same exact-match group, check alternative
                root_j = find(j)
                if root_i != root_j:
                    score = alternative_match_score(pi, pj)
                    if score >= 0.35:
                        alternatives[root_i].append((j, score))
                        alternatives[root_j].append((i, score))

    # Also compare products that weren't matched by brand (generic matching)
    print("  Cross-brand alternative search...")
    # Build token index for cross-brand matching
    token_to_idx = defaultdict(list)
    for i, p in enumerate(all_products):
        for t in p['parsed']['product_key'].split():
            if len(t) > 2:
                token_to_idx[t].append(i)

    # For single-store products, find alternatives via token overlap
    single_store = [i for i in range(n) if find(i) == i]
    count = 0
    for i in single_store:
        pi = all_products[i]['parsed']
        candidates = set()
        for t in pi['product_key'].split():
            if len(t) > 2 and t in token_to_idx:
                for idx in token_to_idx[t]:
                    if idx != i and all_products[idx]['store'] != all_products[i]['store']:
                        candidates.add(idx)

        for j in list(candidates)[:10]:
            pj = all_products[j]['parsed']
            score = alternative_match_score(pi, pj)
            if score >= 0.4:
                alternatives[i].append((j, score))
        count += 1
        if count % 10000 == 0:
            print("  %d/%d single-store checked" % (count, len(single_store)))

    # ── Build groups ──────────────────────────────────────────────────────────
    print("\nBuilding product groups...")
    groups = defaultdict(list)
    for i in range(n):
        root = find(i)
        groups[root].append(i)

    # Build catalog
    catalog = []
    for root, members in groups.items():
        products = [all_products[i] for i in members]
        stores = set(p['store'] for p in products)

        # Best name (shortest = most specific)
        best = min(products, key=lambda p: len(p['parsed']['normalized']))

        # Prices per store
        prices = {}
        links = {}
        imgs = {}
        for p in products:
            prices[p['store']] = p['price']
            links[p['store']] = p['link']
            imgs[p['store']] = p['img']

        # Find alternatives
        alt_list = []
        seen = set()
        for alt_idx, alt_score in alternatives.get(root, []):
            alt_product = all_products[alt_idx]
            alt_root = find(alt_idx)
            if alt_root != root and alt_idx not in seen:
                seen.add(alt_idx)
                alt_list.append({
                    'name': alt_product['name'],
                    'store': alt_product['store'],
                    'price': alt_product['price'],
                    'score': round(alt_score, 3),
                })
        alt_list.sort(key=lambda x: -x['score'])

        catalog.append({
            'name': best['name'],
            'name_ar': best['name_ar'],
            'brand': best['parsed']['brand'],
            'size': '%s %s' % (best['parsed']['size_val'], best['parsed']['size_unit']) if best['parsed']['size_val'] else '',
            'form': best['parsed']['form'],
            'stores': ', '.join(sorted(stores)),
            'stores_count': len(stores),
            'nahdi_price': prices.get('Nahdi', ''),
            'united_price': prices.get('United', ''),
            'aldawaa_price': prices.get('Al-Dawaa', ''),
            'nahdi_link': links.get('Nahdi', ''),
            'united_link': links.get('United', ''),
            'aldawaa_link': links.get('Al-Dawaa', ''),
            'nahdi_img': imgs.get('Nahdi', ''),
            'united_img': imgs.get('United', ''),
            'aldawaa_img': imgs.get('Al-Dawaa', ''),
            'alternatives': alt_list[:5],  # Top 5 alternatives
        })

    catalog.sort(key=lambda x: (-x['stores_count'], x['brand'], x['name']))

    multi = sum(1 for c in catalog if c['stores_count'] >= 2)
    three = sum(1 for c in catalog if c['stores_count'] == 3)
    with_alt = sum(1 for c in catalog if c['alternatives'])
    print("\n=== Results ===")
    print("Unique products: %d" % len(catalog))
    print("In 3 stores: %d" % three)
    print("In 2 stores: %d" % multi)
    print("With alternatives: %d" % with_alt)

    # ── Write Excel ───────────────────────────────────────────────────────────
    print("\nWriting Excel...")
    wb_out = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hf_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    g_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    y_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    r_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    # Sheet 1: Price Comparison (multi-store first)
    ws1 = wb_out.active
    ws1.title = 'Price Comparison'
    cols1 = ['name', 'brand', 'size', 'form', 'stores', 'stores_count',
             'nahdi_price', 'united_price', 'aldawaa_price']
    hdrs1 = ['Product Name', 'Brand', 'Size', 'Form', 'Available In', '# Stores',
             'Nahdi Price', 'United Price', 'Al-Dawaa Price']
    w1 = [55, 25, 15, 15, 25, 12, 14, 14, 14]

    for ci, h in enumerate(hdrs1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hf_fill; c.alignment = Alignment(horizontal='center'); c.border = bd
    for ci, w in enumerate(w1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    for ri, g in enumerate(catalog, 2):
        for ci, col in enumerate(cols1, 1):
            v = g.get(col, '')
            c = ws1.cell(row=ri, column=ci, value=v)
            c.border = bd
            if col == 'stores_count':
                c.fill = g_fill if v >= 3 else (y_fill if v >= 2 else r_fill)
                c.alignment = Alignment(horizontal='center')
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(hdrs1)), len(catalog) + 1)

    # Sheet 2: With Alternatives
    ws2 = wb_out.create_sheet('Alternatives')
    cols2 = ['name', 'brand', 'size', 'stores', 'nahdi_price', 'united_price', 'aldawaa_price',
             'alt_text']
    hdrs2 = ['Product Name', 'Brand', 'Size', 'Stores', 'Nahdi', 'United', 'Al-Dawaa',
             'Alternatives (Store: Name @ Price)']
    w2 = [55, 25, 15, 25, 14, 14, 14, 80]

    for ci, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hf_fill; c.alignment = Alignment(horizontal='center'); c.border = bd
    for ci, w in enumerate(w2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    alt_catalog = [c for c in catalog if c['alternatives']]
    for ri, g in enumerate(alt_catalog, 2):
        for ci, col in enumerate(cols2, 1):
            if col == 'alt_text':
                alt_str = '; '.join(['%s: %s @ %s' % (a['store'], a['name'][:40], a['price']) for a in g['alternatives']])
                v = alt_str
            else:
                v = g.get(col, '')
            c = ws2.cell(row=ri, column=ci, value=v)
            c.border = bd
    ws2.freeze_panes = 'A2'

    # Sheet 3-5: Raw data per store
    for sname, sdata in [('Nahdi', nahdi), ('United', united), ('Al-Dawaa', aldawaa)]:
        ws = wb_out.create_sheet('%s All' % sname)
        hdrs = ['Name (EN)', 'Name (AR)', 'Price (SAR)', 'SKU', 'Brand', 'Link', 'Image']
        cw = [55, 55, 14, 20, 25, 60, 60]
        scols = ['Name (EN)', 'Name (AR)', 'Price (SAR)', 'SKU', 'Brand', 'Product Link', 'Image URL']
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hf; c.fill = hf_fill; c.alignment = Alignment(horizontal='center'); c.border = bd
        for ci, w in enumerate(cw, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri, p in enumerate(sdata, 2):
            for ci, col in enumerate(scols, 1):
                v = p.get(col, '')
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = bd
        ws.freeze_panes = 'A2'

    wb_out.save(OUTPUT)
    print("Saved: %s" % OUTPUT)
    print("Time: %ds" % (time.time() - t0))


if __name__ == '__main__':
    main()
